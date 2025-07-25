

"""
-pandas is a package providing fast,flexible and expressive data structures to make working
with relational or labeled data both easy and intuitive

- it has func for analyzing,cleaning,exploring and manipulating data
- for scientific and mathematical calculation
- also called panel data and python data analysis

applications
-easy handling of missing data
- ssize mutability - col can be inserted and deleted from dataframe and higher dimensional objects

-automatic and explcitdata alignment  means we can create row and col ,and set alignment automatically
 when datas given it auto stores respective row and col

- flexible group by func like pivot table in excel to get data summary and relationships
-intwlligent label-based slicing ,fancy indexing and subsetting of large data sets.

-intuitive merging  and joining data sets
-flexible reshaping and pivoting of data sets- get small data set from big data set
- series is one col with single data type- single dimensional
data frame - series of colmns

panda series is a one dimensional labeled array capable of holding data of any type.axis labels are called index
panda series is nothing but a column in an excel sheet.
supports int and labeled based indexing.

pandas dataframe is two dimensional size mutable ,potentially heterogeneous tabular data structure with
labeled axes(rows and col)
it can be created from csv files,excel,sql data.
using dataframe func we can use data in dictionary,list format
"""
# creation of data frame - 2d data
import pandas as pd
import  numpy as np
data = {"name":["john","doe","lisa","ui"],
        "age":[25,24,0,28],
        "salary":[2000,3000,4000,0]}
df = pd.DataFrame(data)
print(df)
data2  = pd.read_csv("annual-enterprise-survey-2023-financial-year-provisional-size-bands.csv")
data1  = pd.read_csv("industry_sic.csv")
data3  = pd.read_csv("D:/selenium/Sheet1csv.csv")
#data4  = pd.read_excel('file.xlsx')
"""
import openpyxl

dataframe  = openpyxl.load_workbook("file.xlsx") # Define variable to load the dataframe

# Define variable to read sheet
dataframe1 = dataframe.active

# Iterate the loop to read the cell values
for row in range(0, dataframe1.max_row):
    for col in dataframe1.iter_cols(1, dataframe1.max_column):
        print(col[row].value)
"""

# exploreing data
print(data2,"2csv") # default 5 values
#print(data2.head(10))
#print(data2.tail(10))
print(data2.info())
print(data2.describe())
print(data2.isnull().sum())
# print(data.describe()) 'dict' object has no attribute 'describe'
print("----------------------")
# checking duplicates
print(data1)
print(data2["industry_code_ANZSIC"].duplicated())
print(data2["industry_code_ANZSIC"].duplicated().sum())
print(data1["SIC Code"].duplicated())
print(data1["SIC Code"].duplicated().sum())

print(data1.drop_duplicates("SIC Code"))

# null values or missing value - whether to give value, replace(with what values) ,delete

print(data2.isnull().sum())
print(data1.isnull().sum())
print(data2.dropna()) # to delete null values - entire row

print(data2.replace(np.nan,3000))
data2["SIC Code"] = data2["SIC Code"].replace(np.nan,30000) # replace nan= null with 30000 in sic col
print(data2)
print(data2["SIC Code"].mean())  # relace with mean
print(data2.fillna(method= "bfill"))  # relace with below val
print(data2.fillna(method= "ffill"))  # relace with upper val
print(data2.fillna("fill"))




