

"""
like excel
-pandas is a package providing fast,flexible and expressive data structures to make working
with relational or labeled data both easy and intuitive

- it has func for analyzing,cleaning,exploring and manipulating data
- for scientific and mathematical calculation
- also called panel data and python data analysis

applications
-easy handling of missing data
- ssize mutability - col can be inserted and deleted from dataframe and higher dimensional objects

-automatic and explcitdata alignment  means we can create row and col ,and set alignment automatically
 when datas given it auto stores respective row and col

- flexible group by func like pivot table in excel to get data summary and relationships
-intwlligent label-based slicing ,fancy indexing and subsetting of large data sets.

-intuitive merging  and joining data sets
-flexible reshaping and pivoting of data sets- get small data set from big data set

> series is one col with single, any data type- single dimensional labeled array
> data frame - series of colmns- 2dimensional table or grid

panda series is a one dimensional labeled array capable of holding data of any type.axis labels are called index
panda series is nothing but a column in an excel sheet.
supports int and labeled based indexing.

pandas dataframe is two dimensional size mutable ,potentially heterogeneous tabular data structure with
labeled axes(rows and col)
it can be created from csv files,excel,sql data.
using dataframe func we can use data in dictionary,list format
"""
# creation of data frame - 2d data

import pandas as pd
import numpy as np

# series
print(pd.__version__)
lta = [1,2,3,4,5]
series1 = pd.Series(lta, index=["a","b","c","d","e"]) # convert list to series using Series constructor and setting index or labels
series2 = pd.Series(lta, index=["apart 1","apart 2","apart 3","apart 4","apart 5"]) # convert list to series using Series constructor and gives index
# series1 is a object
print(series1)
print(series1.loc["a"]) # returns value at a
print(series1.iloc[0]) # returns value at index 0
series1.loc["a"] = 8  # changes value at a
print(series2)
print(series2[series2 > 2])

# using dict
calories = {"day 1":1500, "day 2": 2000, "day 3": 1700}
calseries = pd.Series(calories)
print(calseries)
print(calseries.loc["day 3"])

calseries.iloc[0] = 1800
calseries.loc["day 3"] += 500
print(calseries)
print(calseries[calseries<2000])

#dataframe - tabular data structure with rows and frames

data ={"name": ["spong","pat","squid"],
       "age": [30,40,50]}

dataframe1 = pd.DataFrame(data)
dataframe2 = pd.DataFrame(data,index=["emp1","emp2","emp3"])

print(dataframe1)
print(dataframe2)
print(dataframe2.loc["emp1"]) # loc means location  by label
print(dataframe2.iloc[1]) # loc means location  by index

# add new col like job,position
dataframe2["job"] = ["cook","na","cashier" ]
print(dataframe2)

#add new row by creating new df and concaneting it

neew_row = pd.DataFrame([{"name":"sandy", "age":34, "job": "eng"},{"name":"sand", "age":3, "job": "en"}],
                        index= ["emp4","emp5"])
dataframe2 = pd.concat([dataframe2, neew_row])
print(dataframe2)

# using csv and json

cs = pd.read_csv("iris-parquet.csv")
cars = pd.read_csv("mtcars-parquet.csv",index_col="model") # name of col to be indexed
jso1 = pd.read_json("house-price-parquet.json",lines= True)

print(cs)
print(jso1)
# print(cs.to_string()) # will print all

# selection by col
print(cs[["sepal.length","petal.width"]])
print("=====================by row")
# selection by row
print(cs.loc[1])  # gives data of 2nd row
print(cars)
# set one of the col to serve as index
print(cars.loc["Merc 280C"])
print(cars.loc["Merc 280C", ["cyl","hp"]]) # only data of cyl and hp
print(cars.loc["Merc 280C":"Merc 450SLC", ["cyl","hp"]]) # only data of cyl and hp

print(cars.iloc[0:11]) # first 10 rows
print(cars.iloc[0:11:2]) # first 10 rows , step of 2
print(cars.iloc[0:11:2, 0:3]) # first 10 rows , step of 2, col 0-3

carn = input("enter a car name: ")
try:
    print(cars.loc[carn])
except KeyError:
    print(f"{carn} not found")

# filteering

powerfull = cars[cars["hp"] >= 200]
print(powerfull)
geared  = cars[cars["gear"] == 5]
print(geared)

pok = cars[(cars["vs"] == 1) | (cars["am"] == 1)]
pok2 = cars[(cars["vs"] == 1) & (cars["am"] == 1)]
print(pok)
print(pok2)
print("------------------------------aggre for whole dataframe")
print(cars.mean(numeric_only= True)) # find mean of every col that is numeric
print(cars.sum(numeric_only= True)) # find sum of every col that is numeric
print(cars.min(numeric_only= True)) # find min of every col that is numeric
print(cars.max(numeric_only= True)) # find max of every col that is numeric
print(cars.count())

print("------------------------------aggre for single col")
print(cars["hp"].mean()) # find mean of every col that is numeric
print(cars["hp"].sum()) # find sum of every col that is numeric
print(cars["gear"].min()) # find min of every col that is numeric
print(cars["vs"].max()) # find max of every col that is numeric
print(cars["am"].count())

# arranging in different groups(object)
group1 = cars.groupby("gear")
print(group1["hp"].mean())
print(group1["hp"].min())
print(group1["hp"].max())
print(group1["hp"].sum())

#data cleaning

# 1 drop a col

cars.drop(columns=["am","vs"])
print(cars)

carrr = cars.drop(columns=["am","vs"])
print(carrr)

# drop missing val
cana = cars.dropna(subset=["gear"]) # rows with na will be droped
print(cana)
# fill blank or na places
cafil = cars.fillna({"gear": "None"})

# fix inconsistent val
cars["gear"] = cars["gear"].replace({3: 6,4:7})
#cars["model"] = cars["model"].replace({"Datsun 710 ": "Datsun 910" })
print(cars)

# standaardize text

cs["variety"] = cs["variety"].str.lower()
print(cs)

# fix data type
cars["vs"] = cars["vs"].astype(bool)
print(cars)

# remove duplicate val

cs = cs.drop_duplicates()
print(cs.to_string())







"""
Explanation:
import pandas as pd: This line imports the pandas library, commonly aliased as pd.
pd.read_json('data.json'): This is the core function call.
It takes the path to your JSON file as an argument (e.g., 'data.json').
It reads the data from the specified JSON file.
It automatically converts the JSON data into a pandas DataFrame.
Handling Different JSON Structures:
JSON Lines (JSONL) format: If your JSON file contains multiple JSON objects, each on a new line (e.g., {"key": "value"}\n{"key": "another_value"}), you need to specify lines=True:


    df = pd.read_json('data_lines.json', lines=True)
Nested JSON: For more complex, nested JSON structures, read_json() might flatten some levels automatically. For deeper nesting or specific flattening requirements, you might need to combine json.loads() (from Python's built-in json module) and pd.json_normalize() for more granular control over the flattening process.


import json
from pandas import json_normalize

# Example with nested data
# data = [{"id": 1, "details": {"name": "Alice", "age": 30}}, ...]
with open('nested_data.json', 'r') as f:
    data = json.load(f)
df = json_normalize(data)
"""
print("---------------------------------------brocode")

data = {"name":["john","doe","lisa","ui"],
        "age":[25,24,0,28],
        "salary":[2000,3000,4000,0]}
df1 = pd.DataFrame(data)
print(df1)
data2  = pd.read_csv("annual-enterprise-survey-2023-financial-year-provisional-size-bands.csv")
data1  = pd.read_csv("industry_sic.csv")
data3  = pd.read_csv("D:/selenium/Sheet1csv.csv")
#data4  = pd.read_excel('file.xlsx')
"""
import openpyxl

dataframe  = openpyxl.load_workbook("file.xlsx") # Define variable to load the dataframe

# Define variable to read sheet
dataframe1 = dataframe.active

# Iterate the loop to read the cell values
for row in range(0, dataframe1.max_row):
    for col in dataframe1.iter_cols(1, dataframe1.max_column):
        print(col[row].value)


# exploreing data
print(data2,"2csv") # default 5 values
#print(data2.head(10))
#print(data2.tail(10))
print(data2.info())
print(data2.describe())
print(data2.isnull().sum())
# print(data.describe()) 'dict' object has no attribute 'describe'
print("----------------------")
# checking duplicates
print(data1)
print(data2["industry_code_ANZSIC"].duplicated())
print(data2["industry_code_ANZSIC"].duplicated().sum())
print(data1["SIC Code"].duplicated())
print(data1["SIC Code"].duplicated().sum())

print(data1.drop_duplicates("SIC Code"))

# null values or missing value - whether to give value, replace(with what values) ,delete

print(data2.isnull().sum())
print(data1.isnull().sum())
print(data2.dropna()) # to delete null values - entire row

print(data2.replace(np.nan,3000))
data2["SIC Code"] = data2["SIC Code"].replace(np.nan,30000) # replace nan= null with 30000 in sic col
print(data2)
print(data2["SIC Code"].mean())  # relace with mean
print(data2.fillna(method= "bfill"))  # relace with below val
print(data2.fillna(method= "ffill"))  # relace with upper val
print(data2.fillna("fill"))

"""


